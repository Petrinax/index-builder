"""
Index Builder Service - Core business logic for index construction and queries
"""

import sys
import os
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Tuple

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from data_pipeline.database import DatabaseFactory, DatabaseConnection
from data_pipeline.base_logging import Logger
from fast_app.models import (
    BuildIndexResponse,
    IndexPerformanceResponse,
    IndexCompositionResponse,
    CompositionChangesResponse,
    DailyPerformance,
    StockComposition,
    CompositionChange,
)
from fast_app.cache import RedisCache

# Setup logger using base_logging module
logger = Logger("fast_app.services")


class IndexBuilderService:
    """Service for building and querying equal-weighted indices"""

    def __init__(self, db_type: str = 'sqlite', db_path: str = None):
        self.db_type = db_type
        self.db_path = db_path or 'data/stock_data.db'

        logger.info(f"Initializing IndexBuilderService with db_type={db_type}, db_path={self.db_path}")

        self._cache = RedisCache()
        logger.info("Redis cache initialized")

        # Create tables for storing index data
        self._init_index_tables()
        logger.info("Index tables initialized")

    def _get_connection(self) -> DatabaseConnection:
        """Get database connection"""
        logger.debug(f"Creating database connection: {self.db_type}")
        return DatabaseFactory.create(self.db_type, self.db_path)

    def _init_index_tables(self):
        """Initialize tables for storing index data"""
        logger.info("Initializing index tables...")
        conn = self._get_connection()

        try:
            # Table for daily index performance
            conn.execute("""
                CREATE TABLE IF NOT EXISTS index_performance (
                    date DATE,
                    nav REAL,
                    daily_return REAL,
                    cumulative_return REAL,
                    top_n INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    PRIMARY KEY (date, top_n)
                )
            """)
            logger.debug("index_performance table created/verified")

            # Table for daily index composition
            conn.execute("""
                CREATE TABLE IF NOT EXISTS index_composition (
                    date DATE,
                    symbol VARCHAR,
                    exchange VARCHAR,
                    market_cap REAL,
                    price REAL,
                    shares REAL,
                    weight REAL,
                    notional_value REAL,
                    top_n INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    PRIMARY KEY (date, symbol, exchange, top_n)
                )
            """)
            logger.debug("index_composition table created/verified")

            # Table for composition changes
            conn.execute("""
                CREATE TABLE IF NOT EXISTS composition_changes (
                    date DATE,
                    symbol VARCHAR,
                    exchange VARCHAR,
                    change_type VARCHAR,
                    market_cap REAL,
                    top_n INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    PRIMARY KEY (date, symbol, exchange, change_type, top_n)
                )
            """)
            logger.debug("composition_changes table created/verified")

        except Exception as e:
            logger.error(f"Error initializing index tables: {str(e)}")
            raise
        finally:
            conn.close()

    def _get_top_n_stocks(self, date: str, top_n: int) -> List[Dict]:
        """Fetch top N stocks by market cap for a given date"""
        logger.debug(f"Fetching top {top_n} stocks for date: {date}")
        conn = self._get_connection()

        try:
            result = conn.execute("""
                SELECT 
                    m.symbol, 
                    m.exchange, 
                    m.market_cap,
                    p.close as price,
                    p.volume
                FROM daily_market_cap m
                JOIN daily_stock_prices p 
                    ON m.symbol = p.symbol 
                    AND m.date = p.date 
                    AND m.exchange = p.exchange
                WHERE m.date = ?
                    AND m.market_cap IS NOT NULL
                    AND p.close IS NOT NULL
                    AND p.close > 0
                ORDER BY m.market_cap DESC
                LIMIT ?
            """, [date, top_n]).fetchall()

            if not result:
                logger.warning(f"No stock data found for date {date}")
                raise ValueError(f"No stock data found for date {date}")

            stocks = []
            for row in result:
                stocks.append({
                    'symbol': row[0],
                    'exchange': row[1],
                    'market_cap': row[2],
                    'price': row[3],
                    'volume': row[4]
                })

            logger.debug(f"Retrieved {len(stocks)} stocks for date {date}")
            return stocks

        except Exception as e:
            logger.error(f"Error fetching top N stocks for {date}: {str(e)}")
            raise
        finally:
            conn.close()

    def _calculate_equal_weights(self, stocks: List[Dict], nav: float) -> List[Dict]:
        """Calculate equal weights and share allocation"""
        logger.debug(f"Calculating equal weights for {len(stocks)} stocks with NAV {nav}")

        n = len(stocks)
        if n == 0:
            logger.warning("No stocks provided for weight calculation")
            return []

        weight = 1.0 / n
        notional_per_stock = nav * weight

        for stock in stocks:
            stock['weight'] = weight
            stock['notional_value'] = notional_per_stock
            stock['shares'] = notional_per_stock / stock['price']

        logger.debug(f"Calculated equal weights: {weight:.4f} per stock, ${notional_per_stock:.2f} notional per stock")
        return stocks

    def _calculate_nav(self, portfolio: List[Dict]) -> float:
        """Calculate current NAV based on portfolio positions"""
        nav = 0.0
        for position in portfolio:
            nav += position['shares'] * position['price']

        logger.debug(f"Calculated NAV: ${nav:.2f} from {len(portfolio)} positions")
        return nav

    def _update_portfolio_prices(self, portfolio: List[Dict], date: str) -> List[Dict]:
        """Update portfolio with current prices for a given date"""
        logger.debug(f"Updating portfolio prices for {len(portfolio)} positions on {date}")
        conn = self._get_connection()

        try:
            updated_portfolio = []
            for position in portfolio:
                result = conn.execute("""
                    SELECT close, volume
                    FROM daily_stock_prices
                    WHERE symbol = ? AND exchange = ? AND date = ?
                """, [position['symbol'], position['exchange'], date]).fetchone()

                outstanding_shares = position['market_cap'] // position['price']

                if result and result[0]:
                    position['price'] = result[0]
                    position['volume'] = result[1]
                    position['notional_value'] = position['shares'] * position['price']
                    position['market_cap'] = outstanding_shares * position['price']
                    updated_portfolio.append(position)
                else:
                    logger.debug(f"No price data for {position['symbol']} on {date}")

            logger.debug(f"Updated {len(updated_portfolio)} positions out of {len(portfolio)} for {date}")
            return updated_portfolio

        except Exception as e:
            logger.error(f"Error updating portfolio prices for {date}: {str(e)}")
            raise
        finally:
            conn.close()

    def _detect_composition_changes(self, old_symbols: set, new_symbols: set,
                                     old_portfolio: List[Dict], new_portfolio: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
        """Detect which stocks were added and removed"""
        added_symbols = new_symbols - old_symbols
        removed_symbols = old_symbols - new_symbols

        added_stocks = [s for s in new_portfolio if (s['symbol'], s['exchange']) in added_symbols]
        removed_stocks = [s for s in old_portfolio if (s['symbol'], s['exchange']) in removed_symbols]

        return added_stocks, removed_stocks

    def _save_performance(self, date: str, nav: float, daily_return: Optional[float],
                          cumulative_return: float, top_n: int):
        """Save performance data to database"""
        logger.info(f"Saving performance data for {date}: NAV={nav:.2f}, Daily={daily_return:.2f}%, Cumulative={cumulative_return:.2f}%")
        conn = self._get_connection()

        try:
            # Delete existing record if any
            conn.execute("""
                DELETE FROM index_performance 
                WHERE date = ? AND top_n = ?
            """, [date, top_n])

            # Insert new record
            conn.execute("""
                INSERT INTO index_performance (date, nav, daily_return, cumulative_return, top_n)
                VALUES (?, ?, ?, ?, ?)
            """, [date, nav, daily_return, cumulative_return, top_n])

            logger.debug(f"Performance data saved for {date}")

        except Exception as e:
            logger.error(f"Error saving performance data for {date}: {str(e)}")
            raise
        finally:
            conn.close()

    def _save_composition(self, date: str, portfolio: List[Dict], top_n: int):
        """Save composition data to database"""
        logger.debug(f"Saving composition data for {date}: {len(portfolio)} positions")
        conn = self._get_connection()

        try:
            # Delete existing records for this date
            conn.execute("""
                DELETE FROM index_composition 
                WHERE date = ? AND top_n = ?
            """, [date, top_n])

            # Insert new records
            for position in portfolio:
                conn.execute("""
                    INSERT INTO index_composition 
                    (date, symbol, exchange, market_cap, price, shares, weight, notional_value, top_n)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, [
                    date,
                    position['symbol'],
                    position['exchange'],
                    position.get('market_cap', 0),
                    position['price'],
                    position['shares'],
                    position['weight'],
                    position['notional_value'],
                    top_n
                ])

            logger.debug(f"Composition data saved for {date}")

        except Exception as e:
            logger.error(f"Error saving composition data for {date}: {str(e)}")
            raise
        finally:
            conn.close()

    def _save_composition_changes(self, date: str, added: List[Dict], removed: List[Dict], top_n: int):
        """Save composition changes to database"""
        logger.debug(f"Saving composition changes for {date}: {len(added)} added, {len(removed)} removed")
        conn = self._get_connection()

        try:
            # Delete existing records
            conn.execute("""
                DELETE FROM composition_changes 
                WHERE date = ? AND top_n = ?
            """, [date, top_n])

            # Insert added stocks
            for stock in added:
                conn.execute("""
                    INSERT INTO composition_changes 
                    (date, symbol, exchange, change_type, market_cap, top_n)
                    VALUES (?, ?, ?, 'ADDED', ?, ?)
                """, [date, stock['symbol'], stock['exchange'], stock.get('market_cap', 0), top_n])

            # Insert removed stocks
            for stock in removed:
                conn.execute("""
                    INSERT INTO composition_changes 
                    (date, symbol, exchange, change_type, market_cap, top_n)
                    VALUES (?, ?, ?, 'REMOVED', ?, ?)
                """, [date, stock['symbol'], stock['exchange'], stock.get('market_cap', 0), top_n])

            logger.debug(f"Composition changes saved for {date}")

        except Exception as e:
            logger.error(f"Error saving composition changes for {date}: {str(e)}")
            raise
        finally:
            conn.close()

    async def build_index(self, start_date: str, end_date: Optional[str],
                          top_n: int, initial_nav: float) -> BuildIndexResponse:
        """
        Build index following the algorithm in build_steps.txt
        """

        logger.info(f"Starting index build: {start_date} to {end_date}, top_n={top_n}, initial_nav={initial_nav}")

        logger.info("Resetting existing index data...")
        self.reset_index_tables()

        # Validate dates
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        if end_date:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        else:
            end_dt = datetime.now()

        logger.info(f"Date range validated: {start_dt.date()} to {end_dt.date()}")

        # INIT Phase: Get data for (start - 1)
        init_date = (start_dt - timedelta(days=1)).strftime('%Y-%m-%d')
        logger.info(f"INIT Phase: Looking for initial data on {init_date}")

        # Try to find a valid trading day before start_date
        attempts = 0
        max_attempts = 10
        top_stocks = []  # Initialize to avoid potential reference before assignment

        while attempts < max_attempts:
            try:
                top_stocks = self._get_top_n_stocks(init_date, top_n)
                logger.info(f"Found initial data on {init_date} after {attempts + 1} attempts")
                break
            except ValueError:
                init_date = (datetime.strptime(init_date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
                attempts += 1
                logger.debug(f"Attempt {attempts}: No data on {init_date}, trying previous day")

        if attempts == max_attempts or not top_stocks:
            logger.error(f"Could not find valid trading data before {start_date} after {max_attempts} attempts")
            raise ValueError(f"Could not find valid trading data before {start_date}")

        # Step 1: Set initial NAV
        current_nav = initial_nav
        base_nav = initial_nav
        logger.info(f"Initial NAV set to ${initial_nav}")

        # Step 2-3: Fetch top N stocks and distribute equally weighted
        portfolio = self._calculate_equal_weights(top_stocks, current_nav)
        prev_portfolio = portfolio.copy()

        # Save initial composition
        self._save_composition(init_date, portfolio, top_n)
        logger.info(f"Initial composition saved for {init_date}")

        # ITERATE Phase
        current_date = start_dt
        days_processed = 0
        logger.info("Starting ITERATE phase...")

        while current_date <= end_dt:
            date_str = current_date.strftime('%Y-%m-%d')
            logger.debug(f"Processing date: {date_str}")

            try:
                # Step 5: Fetch portfolio from previous day
                # Step 6: Update prices and calculate NAV at EOD
                portfolio = self._update_portfolio_prices(prev_portfolio, date_str)

                if not portfolio:
                    # Skip non-trading days
                    logger.info(f"No trading data for {date_str}, skipping")
                    current_date += timedelta(days=1)
                    continue

                prev_nav = current_nav
                current_nav = self._calculate_nav(portfolio)

                # Step 7-8: Calculate returns
                daily_return = ((current_nav - prev_nav) / prev_nav) * 100 if prev_nav > 0 else 0
                cumulative_return = ((current_nav - base_nav) / base_nav) * 100

                logger.debug(f"{date_str}: NAV=${current_nav:.2f}, Daily={daily_return:.2f}%, Cumulative={cumulative_return:.2f}%")

                # Save performance
                self._save_performance(date_str, current_nav, daily_return, cumulative_return, top_n)

                # Step 9: Fetch top N stocks for reconstitution
                new_top_stocks = self._get_top_n_stocks(date_str, top_n)

                # Step 10: Detect composition changes
                old_symbols = {(s['symbol'], s['exchange']) for s in portfolio}
                new_symbols = {(s['symbol'], s['exchange']) for s in new_top_stocks}

                added_stocks, removed_stocks = self._detect_composition_changes(
                    old_symbols, new_symbols, portfolio, new_top_stocks
                )

                # Save composition changes if any
                if added_stocks or removed_stocks:
                    logger.info(f"Composition changes detected on {date_str} - Added: {len(added_stocks)}, Removed: {len(removed_stocks)}")

                    self._save_composition_changes(date_str, added_stocks, removed_stocks, top_n)

                    if added_stocks:
                        logger.debug(f"Added stocks: {[(s['symbol'], s['exchange']) for s in added_stocks]}")
                    if removed_stocks:
                        logger.debug(f"Removed stocks: {[(s['symbol'], s['exchange']) for s in removed_stocks]}")

                # Step 11: Rebalance portfolio with equal weighting
                portfolio = self._calculate_equal_weights(new_top_stocks, current_nav)

                # Step 12: Save composition
                self._save_composition(date_str, portfolio, top_n)

                prev_portfolio = portfolio.copy()
                days_processed += 1

                if days_processed % 7 == 0:  # Log progress every 7 days
                    logger.info(f"Progress: Processed {days_processed} days, current NAV: ${current_nav:.2f}")

            except ValueError as e:
                # Skip days with no data
                logger.debug(f"Skipping {date_str}: {str(e)}")
                pass

            current_date += timedelta(days=1)

        total_return = ((current_nav - base_nav) / base_nav) * 100

        logger.info(f"Index build completed - Days processed: {days_processed}, Final NAV: ${current_nav:.2f}, Total return: {total_return:.2f}%")

        return BuildIndexResponse(
            message="Index built successfully",
            start_date=start_date,
            end_date=end_dt.strftime('%Y-%m-%d'),
            top_n=top_n,
            initial_nav=initial_nav,
            final_nav=current_nav,
            total_return=total_return,
            days_processed=days_processed
        )

    async def get_performance(self, start_date: str, end_date: Optional[str]) -> IndexPerformanceResponse:
        """Get performance data for date range (cached)"""
        if not end_date:
            end_date = datetime.now().strftime("%Y-%m-%d")
        cache_key = f"perf_{start_date}_{end_date}"

        logger.debug(f"Fetching performance data: {start_date} to {end_date}")

        # Check cache first
        cached = self._cache.get(cache_key)
        if cached:
            logger.info(f"Performance data retrieved from cache for {start_date} to {end_date}")
            return IndexPerformanceResponse(**cached)

        logger.debug("Performance data not in cache, querying database")
        conn = self._get_connection()

        try:
            if end_date:
                result = conn.execute("""
                    SELECT date, nav, daily_return, cumulative_return
                    FROM index_performance
                    WHERE date >= ? AND date <= ?
                    ORDER BY date
                """, [start_date, end_date]).fetchall()
            else:
                result = conn.execute("""
                    SELECT date, nav, daily_return, cumulative_return
                    FROM index_performance
                    WHERE date = ?
                """, [start_date]).fetchall()

            if not result:
                logger.warning(f"No performance data found for {start_date} to {end_date}")
                raise ValueError("No performance data found for the specified date range")

            performance = [
                DailyPerformance(
                    date=row[0],
                    nav=row[1],
                    daily_return=row[2],
                    cumulative_return=row[3]
                )
                for row in result
            ]

            # Calculate summary statistics
            summary = {
                "total_days": len(performance),
                "starting_nav": performance[0].nav,
                "ending_nav": performance[-1].nav,
                "total_return": performance[-1].cumulative_return,
                "avg_daily_return": sum(p.daily_return for p in performance if p.daily_return) / len(performance),
                "max_daily_return": max((p.daily_return for p in performance if p.daily_return), default=0),
                "min_daily_return": min((p.daily_return for p in performance if p.daily_return), default=0)
            }

            response = IndexPerformanceResponse(
                start_date=start_date,
                end_date=end_date or start_date,
                performance=performance,
                summary=summary
            )

            # Cache the response
            self._cache.set(cache_key, response.dict())
            logger.info(f"Performance data cached and returned for {start_date} to {end_date}")

            return response

        except Exception as e:
            logger.error(f"Error fetching performance data: {str(e)}")
            raise
        finally:
            conn.close()

    async def get_composition(self, date: str) -> IndexCompositionResponse:
        """Get index composition for a specific date (cached)"""
        cache_key = f"comp_{date}"

        logger.debug(f"Fetching composition data for {date}")

        # Check cache first
        cached = self._cache.get(cache_key)
        if cached:
            logger.info(f"Composition data retrieved from cache for {date}")
            return IndexCompositionResponse(**cached)

        logger.debug("Composition data not in cache, querying database")
        conn = self._get_connection()

        try:
            result = conn.execute("""
                SELECT symbol, exchange, market_cap, price, shares, weight, notional_value
                FROM index_composition
                WHERE date = ?
                ORDER BY market_cap DESC
            """, [date]).fetchall()

            if not result:
                logger.warning(f"No composition data found for {date}")
                raise ValueError(f"No composition data found for date {date}")

            composition = [
                StockComposition(
                    symbol=row[0],
                    exchange=row[1],
                    market_cap=row[2],
                    price=row[3],
                    shares=row[4],
                    weight=row[5],
                    notional_value=row[6]
                )
                for row in result
            ]

            total_market_cap = sum(s.market_cap for s in composition)

            response = IndexCompositionResponse(
                date=date,
                composition=composition,
                total_stocks=len(composition),
                total_market_cap=total_market_cap
            )

            # Cache the response
            self._cache.set(cache_key, response.dict())
            logger.info(f"Composition data cached and returned for {date}")

            return response

        except Exception as e:
            logger.error(f"Error fetching composition data: {str(e)}")
            raise
        finally:
            conn.close()

    async def get_composition_changes(self, start_date: str, end_date: Optional[str]) -> CompositionChangesResponse:
        """Get composition changes for date range (cached)"""
        if not end_date:
            end_date = datetime.now().strftime("%Y-%m-%d")
        cache_key = f"changes_{start_date}_{end_date}"

        logger.debug(f"Fetching composition changes: {start_date} to {end_date}")

        # Check cache first
        cached = self._cache.get(cache_key)
        if cached:
            logger.info(f"Composition changes retrieved from cache for {start_date} to {end_date}")
            return CompositionChangesResponse(**cached)

        logger.debug("Composition changes not in cache, querying database")
        conn = self._get_connection()

        try:
            if end_date:
                result = conn.execute("""
                    SELECT date, symbol, exchange, change_type, market_cap
                    FROM composition_changes
                    WHERE date >= ? AND date <= ?
                    ORDER BY date, change_type, market_cap DESC
                """, [start_date, end_date]).fetchall()
            else:
                result = conn.execute("""
                    SELECT date, symbol, exchange, change_type, market_cap
                    FROM composition_changes
                    WHERE date = ?
                    ORDER BY change_type, market_cap DESC
                """, [start_date]).fetchall()

            # Group by date
            changes_by_date = {}
            for row in result:
                date = row[0]
                if date not in changes_by_date:
                    changes_by_date[date] = {'added': [], 'removed': []}

                stock_info = {
                    'symbol': row[1],
                    'exchange': row[2],
                    'market_cap': row[4]
                }

                if row[3] == 'ADDED':
                    changes_by_date[date]['added'].append(stock_info)
                else:
                    changes_by_date[date]['removed'].append(stock_info)

            changes = [
                CompositionChange(
                    date=date,
                    stocks_added=data['added'],
                    stocks_removed=data['removed'],
                    num_added=len(data['added']),
                    num_removed=len(data['removed'])
                )
                for date, data in sorted(changes_by_date.items())
            ]

            response = CompositionChangesResponse(
                start_date=start_date,
                end_date=end_date or start_date,
                changes=changes,
                total_change_days=len(changes)
            )

            # Cache the response
            self._cache.set(cache_key, response.dict())
            logger.info(f"Composition changes cached and returned for {start_date} to {end_date}")

            return response

        except Exception as e:
            logger.error(f"Error fetching composition changes: {str(e)}")
            raise
        finally:
            conn.close()

    async def export_to_excel(self, start_date: str, end_date: Optional[str]) -> str:
        """Export all data to Excel file"""
        logger.info(f"Starting Excel export for {start_date} to {end_date}")

        from fast_app.config import settings
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows

        try:
            # Get all data
            logger.debug("Fetching performance data for export")
            performance = await self.get_performance(start_date, end_date)

            logger.debug("Fetching composition changes for export")
            changes = await self.get_composition_changes(start_date, end_date)

            # Create Excel workbook
            logger.debug("Creating Excel workbook")
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Sheet 1: Performance
            ws_perf = wb.create_sheet("Performance")
            ws_perf.append(["Date", "NAV", "Daily Return (%)", "Cumulative Return (%)"])

            # Style header
            for cell in ws_perf[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            for p in performance.performance:
                ws_perf.append([p.date, p.nav, p.daily_return, p.cumulative_return])

            # Sheet 2: Summary
            ws_summary = wb.create_sheet("Summary")
            ws_summary.append(["Metric", "Value"])
            for cell in ws_summary[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            for key, value in performance.summary.items():
                ws_summary.append([key.replace('_', ' ').title(), value])

            # Sheet 3: Composition Changes
            ws_changes = wb.create_sheet("Composition Changes")
            ws_changes.append(["Date", "Type", "Symbol", "Exchange", "Market Cap"])

            for cell in ws_changes[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            for change in changes.changes:
                for stock in change.stocks_added:
                    ws_changes.append([change.date, "ADDED", stock['symbol'], stock['exchange'], stock['market_cap']])
                for stock in change.stocks_removed:
                    ws_changes.append([change.date, "REMOVED", stock['symbol'], stock['exchange'], stock['market_cap']])

            # Auto-size columns
            logger.debug("Formatting Excel sheets")
            for ws in [ws_perf, ws_summary, ws_changes]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"index_export_{start_date}_{end_date or start_date}_{timestamp}.xlsx"
            filepath = os.path.join(settings.EXPORT_DIR, filename)

            wb.save(filepath)
            logger.info(f"Excel export completed successfully: {filename}")
            return filepath

        except Exception as e:
            logger.error(f"Error during Excel export: {str(e)}")
            raise

    def clear_cache(self):
        """Clear all cached data"""
        logger.info("Clearing all cache data")
        try:
            self._cache.clear_all()
            logger.info("Cache cleared successfully")
        except Exception as e:
            logger.error(f"Error clearing cache: {str(e)}")
            raise

    def reset_index_tables(self):
        """Reset index-related database tables"""
        logger.warning("Resetting index tables - all index data will be deleted")
        conn = self._get_connection()

        try:
            conn.execute("DELETE FROM index_performance WHERE true")
            conn.execute("DELETE FROM index_composition WHERE true")
            conn.execute("DELETE FROM composition_changes WHERE true")
            logger.info("Index tables cleared")

            logger.info("Index tables reset completed successfully")

        except Exception as e:
            logger.error(f"Error resetting index tables: {str(e)}")
            raise
        finally:
            conn.close()

    def reset_database(self):
        """Reset index-related database tables"""
        logger.warning("Resetting database - all index data will be deleted")
        conn = self._get_connection()

        try:
            conn.execute("DROP TABLE IF EXISTS index_performance")
            conn.execute("DROP TABLE IF EXISTS index_composition")
            conn.execute("DROP TABLE IF EXISTS composition_changes")
            logger.info("Index tables dropped")

            # Re-initialize tables
            self._init_index_tables()
            logger.info("Database reset completed successfully")

        except Exception as e:
            logger.error(f"Error resetting database: {str(e)}")
            raise
        finally:
            conn.close()
